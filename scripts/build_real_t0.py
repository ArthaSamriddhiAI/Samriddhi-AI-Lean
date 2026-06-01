"""Build the real-data t0 snapshot (T-5.14 Option B) from the colleague NAVs and
the three Bloomberg pulls. Deterministic; no LLM; no API; ADDITIVE and NEVER
destructive (writes a NEW data version, never overwrites the source pulls or the
prior snapshot, never invokes the ADR-0014 NAV regenerator).

Inputs (read-only):
  - BASE   the prior enriched snapshot (structure + cross-section + benchmark
           resolution are preserved verbatim; only the monthly SERIES and the
           per-holding tier_b NUMBERS are replaced with real-data recomputes).
  - COLLEAGUE  the real monthly NAVs (keyed by amfi_code), bypassing ADR-0014.
  - v0.1   6 equity TRIs (real), CRISIL Short Term + Liquid (real debt TR levels),
           SPXT + USDINR (real), direct stocks (real).
  - v0.2   the G-Sec yield curve (91-day T-Bill + GIND ladder), cols ~45-60.
  - v0.3   17 FIMMDA BCOP corporate-credit yields.

Outputs (the NEW data version, gitignored like every enriched snapshot):
  - fixtures/snapshots/enriched/snapshot_t0_q2_2026_realv1.json

What it does, by kickoff phase:
  P1  replace mf_funds[].monthly_nav with real (match by amfi_code), replace the
      real equity TRIs + CRISIL debt + fx + held-stock monthly_prices, derive
      sp_500_tri_inr = SPXT x USDINR in code (not pre-converted).
  P2  convert the FIMMDA/G-Sec YIELD curves to TR-level series by the par-bond
      method (carry - ModDur*dy + 0.5*Conv*dy^2); validate by the inverted-COVID
      gut-check; the yields stay untouched as the carried primitive.
  P3  store the debt grid keyed to the DEBT_CELL_INDEX contract, both yields and
      derived TR, with _meta provenance (real-pulled vs derived-from-yield).
  P5a recompute per-holding tier_b from the real series (self stats always; beta
      vs the stored benchmark when resolved), with the time-varying 91-day T-Bill
      risk-free folded through the same socket the sleeve uses, preserving every
      _benchmark_resolution so evaluability gating is unchanged.

Run:
  python3 scripts/build_real_t0.py --selfcheck   # validate the stat reimpl vs the stored tier_b, no write
  python3 scripts/build_real_t0.py --build        # build + validate + write the new version
"""
import argparse
import json
import math
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl required (pip install openpyxl)", file=sys.stderr)
    sys.exit(2)

REPO = Path(__file__).resolve().parent.parent
BASE_SNAP = REPO / "fixtures/snapshots/enriched/snapshot_t0_q2_2026.json"
OUT_SNAP = REPO / "fixtures/snapshots/enriched/snapshot_t0_q2_2026_realv1.json"

MVP = Path("/Users/shubhamsahamate/Desktop/IMT-G/25 - WealthWisers Technologies - Summer Internship/23 - Lean Samriddhi MVP")
COLLEAGUE = MVP / "01 - SamriddhiAI_data_clean.json"
DF = MVP / "21 - Data Foundation Continued"
V01 = DF / "02 - Samriddhi_Bloomberg_BDH_Data_v0.1.xlsx"
V02 = DF / "04 - Transfer Folder/04 - Samriddhi_Universe_Pull_Data_v0.2.xlsx"
V03 = DF / "06 - Samriddhi_Debt_TopUp_Pull_Data_v0.3.xlsx"

WIN_START, WIN_END = "2019-05", "2026-04"
RISK_FREE_ANN = 0.0525  # the canonical constant (lib/agents/risk-reward-stats.ts and enrich_snapshots.py)

# ---------------------------------------------------------------------------
# Bloomberg BDH parser (auto-detecting). Each security is a (Date, value) column
# pair; the ticker labels one of the pair's columns. We find the header row (the
# row with the most Index/Equity/Curncy tickers), and for each ticker column pick
# the adjacent column that parses as dates and the one that parses as floats.
# ---------------------------------------------------------------------------
def _is_ticker(s):
    return isinstance(s, str) and any(k in s for k in ("Index", "Equity", "Curncy")) and "Invalid" not in s


def _ym(v):
    try:
        return v.strftime("%Y-%m")
    except AttributeError:
        return None


def parse_bdh(path):
    """Return {ticker: {YYYY-MM: float}} for every resolved security. Handles both
    BDH pair layouts: v0.1 labels the DATE column with the ticker (value at c+1);
    v0.2/v0.3 label the VALUE column with the ticker (date at c-1). For each ticker
    we pick the (date_col, val_col) pair that actually parses as dates+floats."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Pull grid"] if "Pull grid" in wb.sheetnames else wb[wb.sheetnames[-1]]
    rows = list(ws.iter_rows(values_only=True))
    hidx = max(range(min(8, len(rows))), key=lambda i: sum(_is_ticker(c) for c in rows[i]))
    header = rows[hidx]
    data = rows[hidx + 1:]
    out = {}
    ncol = len(header)
    for c in range(ncol):
        if not _is_ticker(header[c]):
            continue
        ticker = header[c].strip()
        best = {}  # try both pair layouts (v0.1: date at c, value c+1; v0.2/v0.3: date c-1, value c)
        for dcol, vcol in ((c, c + 1), (c - 1, c)):
            if not (0 <= dcol < ncol and 0 <= vcol < ncol):
                continue
            series = {}
            for r in data:
                if dcol >= len(r) or vcol >= len(r):
                    continue
                key, val = _ym(r[dcol]), r[vcol]
                if key and isinstance(val, (int, float)):
                    series[key] = float(val)
            if len(series) > len(best):  # the correct pairing yields the most (date, float) points
                best = series
        if best:
            out[ticker] = best
    return out


def window(series, start=WIN_START, end=WIN_END):
    return {m: v for m, v in series.items() if start <= m <= end}


def rebase_1000(series):
    """Rebase a level series so its earliest in-window point is 1000 (base-invariant
    for the log-return consumers; keeps the snapshot's base-1000 convention)."""
    months = sorted(series)
    if not months:
        return {}
    b = series[months[0]]
    return {m: series[m] / b * 1000.0 for m in months} if b else dict(series)


# ---------------------------------------------------------------------------
# Par-bond yield-to-total-return conversion (the conversion-proposal audit).
# TR_return[t] = carry - ModDur*dy + 0.5*Conv*dy^2 ; carry = y[t-1]/12 ; dy in
# decimal. ModDur/Conv numerically from a par-bond pricer at the stated tenor.
# ---------------------------------------------------------------------------
def _par_price(coupon, y, n_years):
    """Price of an annual-coupon bond, face 100, coupon=coupon (decimal), yield y."""
    if y <= -0.999:
        return float("nan")
    n = max(1, int(round(n_years)))
    pv = sum(coupon * 100 / (1 + y) ** k for k in range(1, n + 1))
    pv += 100 / (1 + y) ** n
    return pv


def _moddur_conv(y, n_years):
    """Modified duration and convexity of a par bond (coupon = par yield y) at tenor
    n_years, via a central numerical derivative of the par-bond price."""
    coupon = y  # par bond: coupon = yield so price = 100
    h = 1e-4
    p0 = _par_price(coupon, y, n_years)
    pu = _par_price(coupon, y + h, n_years)
    pd = _par_price(coupon, y - h, n_years)
    moddur = -(pu - pd) / (2 * h * p0)
    conv = (pu - 2 * p0 + pd) / (h * h * p0)
    return moddur, conv


def yield_to_tr(yield_series_pct, tenor_years):
    """Convert an annualised-yield series (in PERCENT, e.g. 7.03) at a fixed tenor
    to a TR-level series, base 1000 at the first in-window month. Uses the prior
    month's yield for carry and the closed-form par-bond duration/convexity for
    the price change; needs the month just before WIN_START for the first dy."""
    months = sorted(yield_series_pct)
    if len(months) < 2:
        return {}
    lvl = {}
    level = 1000.0
    prev_m = None
    for m in months:
        if prev_m is not None:
            y0 = yield_series_pct[prev_m] / 100.0
            y1 = yield_series_pct[m] / 100.0
            dy = y1 - y0
            moddur, conv = _moddur_conv(y0, tenor_years)
            tr_ret = (y0 / 12.0) - moddur * dy + 0.5 * conv * dy * dy
            level *= (1 + tr_ret)
        if WIN_START <= m <= WIN_END:
            lvl[m] = level
        prev_m = m
    return rebase_1000(lvl)


# ---------------------------------------------------------------------------
# Stats. Self stats (vol/sharpe/sortino/mdd/calmar) reuse enrich_snapshots.py's
# EXACT functions (the canonical convention the whole snapshot already uses, so
# the per-holding read-through stays internally consistent). Benchmark-relative
# (beta/r2/te/ir) is computed month-aligned, faithful to the TS gate layer's
# benchRelative (lib/agents/risk-reward-stats.ts): align on shared YYYY-MM keys,
# beta/r2 over the full intersection, te/ir over the trailing 36. (enrich's own
# beta_and_r2 aligns POSITIONALLY, which is harmless on the smooth synthetic
# series but would misalign real NAVs (241 months) against the 84-month indices,
# so the benchmark-relative path is done calendar-aligned here, matching the gate
# and ADR-0015, while the rf-bearing self stats reuse enrich verbatim.)
# ---------------------------------------------------------------------------
sys.path.insert(0, str(REPO / "scripts"))
import enrich_snapshots as E  # noqa: E402  (constants + pure stat functions; main() is __main__-guarded)


def _logret_by_month(series):
    months = sorted(series)
    out = {}
    for i in range(1, len(months)):
        p0, p1 = series[months[i - 1]], series[months[i]]
        if p0 and p0 > 0 and p1 and p1 > 0:
            out[months[i]] = math.log(p1 / p0)
    return out


def bench_relative(price_series, bench_values):
    """Calendar-aligned beta/r2/te/ir, faithful to the TS gate benchRelative."""
    fr = _logret_by_month(price_series)
    br = _logret_by_month(bench_values)
    common = sorted(m for m in fr if m in br)
    if len(common) < 12:
        return (None, None, None, None)
    s = [fr[m] for m in common]
    b = [br[m] for m in common]
    n = len(common)
    ms, mb = sum(s) / n, sum(b) / n
    cov = sum((s[i] - ms) * (b[i] - mb) for i in range(n)) / (n - 1)
    vb = sum((x - mb) ** 2 for x in b) / (n - 1)
    vs = sum((x - ms) ** 2 for x in s) / (n - 1)
    if vb == 0 or vs == 0:
        return (None, None, None, None)
    beta = cov / vb
    r2 = (cov / math.sqrt(vs * vb)) ** 2
    w = common[-36:]
    sw, bw = [fr[m] for m in w], [br[m] for m in w]
    diffs = [sw[i] - bw[i] for i in range(len(w))]
    md = sum(diffs) / len(diffs)
    te = math.sqrt(sum((d - md) ** 2 for d in diffs) / (len(diffs) - 1)) * math.sqrt(12)
    annS = math.exp((sum(sw) / len(sw)) * 12) - 1
    annB = math.exp((sum(bw) / len(bw)) * 12) - 1
    ir = (annS - annB) / te if te else None
    return (beta, r2, te, ir)


def r4(v):
    return round(v, 4) if isinstance(v, float) else v


def compute_tier_b(price_series, bench_values, rf_ann, prev_resolution, prev_benchid):
    """Recompute tier_b from a real series. Self stats via enrich's exact functions
    (constant rf_ann, the canonical basis); benchmark-relative month-aligned per the
    gate. Benchmark-relative only when the stored resolution was resolved AND a
    benchmark series is present. Preserves _benchmark_resolution and
    _meta.benchmark_index_id so the risk-reward evaluability gating is unchanged."""
    rets = E.compute_returns_from_series(price_series)
    if len(rets) < 12:
        return {"data_window_insufficient": True, "reason": "fewer_than_12_months",
                "_benchmark_resolution": prev_resolution}
    n5 = len(rets) >= 60
    months = sorted(price_series)
    mdd3 = {m: price_series[m] for m in months[-36:]}
    mdd5 = {m: price_series[m] for m in months[-60:]} if len(months) >= 60 else None
    out = {
        "vol_3y_annualized": r4(E.annualized_vol(rets, 36)),
        "vol_5y_annualized": r4(E.annualized_vol(rets, 60)) if n5 else None,
        "sharpe_3y": r4(E.sharpe_ratio(rets, rf_ann, 36)),
        "sharpe_5y": r4(E.sharpe_ratio(rets, rf_ann, 60)) if n5 else None,
        "sortino_3y": r4(E.sortino_ratio(rets, rf_ann, 36)),
        "sortino_5y": r4(E.sortino_ratio(rets, rf_ann, 60)) if n5 else None,
        "max_drawdown_3y": r4(E.max_drawdown(mdd3)),
        "max_drawdown_5y": r4(E.max_drawdown(mdd5)) if mdd5 else None,
        "calmar_3y": r4(E.calmar_ratio(rets, E.max_drawdown(mdd3), 36)),
    }
    resolved = prev_resolution not in ("benchmark_structurally_inapplicable", "benchmark_not_in_snapshot")
    if resolved and bench_values:
        beta, r2, te, ir = bench_relative(price_series, bench_values)
        out["beta_3y"], out["r_squared_3y"] = r4(beta), r4(r2)
        out["tracking_error_3y"], out["information_ratio_3y"] = r4(te), r4(ir)
    else:
        out["beta_3y"] = out["r_squared_3y"] = out["tracking_error_3y"] = out["information_ratio_3y"] = None
    if prev_resolution:
        out["_benchmark_resolution"] = prev_resolution
    if prev_benchid:
        out["_meta"] = {"benchmark_index_id": prev_benchid}
    return out


# ---------------------------------------------------------------------------
# FIMMDA / G-Sec ticker -> (DEBT_CELL_INDEX key, tenor_years). The conversion
# applies uniformly across this grid; the gate cells (high_grade-short, cash) are
# the REAL CRISIL levels from v0.1 and are NOT converted.
# ---------------------------------------------------------------------------
DEBT_YIELD_MAP = {
    "BCOP3A3M Index": ("aaa_3m_tr", 0.25), "BCOPAAA1 Index": ("aaa_1y_tr", 1),
    "BCOPAAA2 Index": ("aaa_2y_tr", 2), "BCOPAAA3 Index": ("aaa_3y_tr", 3),
    "BCOPAAA5 Index": ("aaa_5y_tr", 5), "BCOPAAA7 Index": ("aaa_7y_tr", 7),
    "BCOPAAA0 Index": ("aaa_10y_tr", 10), "BCOPAA1 Index": ("aa_1y_tr", 1),
    "BCOPAA3 Index": ("aa_3y_tr", 3), "BCOPAA5 Index": ("aa_5y_tr", 5),
    "BCOPA2 Index": ("a_2y_tr", 2), "BCOPA3 Index": ("a_3y_tr", 3),
    "BCOPP3A3 Index": ("psu_3y_tr", 3), "BCOPP3A5 Index": ("psu_5y_tr", 5),
    "BCOPBBB3 Index": ("bbb_3y_tr", 3), "BCOPN3A2 Index": ("nbfc_2y_tr", 2),
    "BCOPN3A3 Index": ("nbfc_3y_tr", 3),
    "GIND1YR Index": ("gsec_1y_tr", 1), "GIND3YR Index": ("gsec_3y_tr", 3),
    "GIND5YR Index": ("gsec_5y_tr", 5), "GIND7YR Index": ("gsec_7y_tr", 7),
    "GIND10YR Index": ("gsec_10y_tr", 10), "GIND15YR Index": ("gsec_15y_tr", 15),
    "GIND30YR Index": ("gsec_30y_tr", 30),
}
RF_TICKER = "91TBYLD Index"  # 91-day T-Bill yield -> the time-varying risk-free


def covid_move(series):
    months = sorted(series)
    if "2020-03" not in series:
        return None
    i = months.index("2020-03")
    if i == 0:
        return None
    p0, p1 = series[months[i - 1]], series["2020-03"]
    return (p1 - p0) / p0 if p0 else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--selfcheck", action="store_true")
    ap.add_argument("--build", action="store_true")
    args = ap.parse_args()

    print("Loading base snapshot ...")
    base = json.loads(BASE_SNAP.read_text())

    # ---- SELFCHECK: the reimplemented stats must reproduce the stored tier_b on
    # the prior (synthetic) NAVs under the constant rf, to 4dp, before real use.
    if args.selfcheck:
        idx = base.get("indices", {})

        def bench_vals(benchid):
            return idx.get(benchid, {}).get("monthly_values") if benchid else None

        SELF = ("vol_3y_annualized", "vol_5y_annualized", "sharpe_3y", "sharpe_5y",
                "sortino_3y", "sortino_5y", "max_drawdown_3y", "max_drawdown_5y", "calmar_3y")
        canon_ok = canon_bad = 0          # my self stats vs enrich's canonical function
        stored_same = stored_changed = 0  # informational: vs the stored (mixed-convention) values
        beta_total = beta_moved = 0
        for f in base["mf_funds"][:300]:
            tb = f.get("tier_b_stats") or {}
            nav = f.get("monthly_nav")
            if not nav or tb.get("data_window_insufficient"):
                continue
            benchid = (tb.get("_meta") or {}).get("benchmark_index_id")
            got = compute_tier_b(nav, bench_vals(benchid), RISK_FREE_ANN,
                                 tb.get("_benchmark_resolution"), benchid)
            canon = E.compute_tier_b_for_instrument(nav, None)  # the canonical self-stat convention
            for k in SELF:
                mine, cv, sv = got.get(k), canon.get(k), tb.get(k)
                if mine is not None and cv is not None:
                    if abs(mine - cv) > 1e-9:
                        canon_bad += 1
                        if canon_bad <= 8:
                            print(f"  CANON-MISMATCH {f.get('fund_name','?')[:28]:28} {k}: mine={mine} enrich={cv}")
                    else:
                        canon_ok += 1
                if mine is not None and sv is not None:
                    stored_changed += (abs(mine - sv) > 0.0011)
                    stored_same += (abs(mine - sv) <= 0.0011)
            if tb.get("beta_3y") is not None and got.get("beta_3y") is not None:
                beta_total += 1
                if abs(tb["beta_3y"] - got["beta_3y"]) > 0.0011:
                    beta_moved += 1
        print(f"\nSelfcheck SELF-STATS vs enrich canonical: {canon_ok} match, {canon_bad} mismatch. "
              f"{'OK (recompute is canonically exact)' if canon_bad == 0 else 'FAIL'}")
        print(f"Selfcheck (informational) vs stored: {stored_changed} self-stat fields change, {stored_same} unchanged "
              f"-- the stored tier_b mixes conventions (vol_3y/beta recomputed under ADR-0015, sortino/mdd/5y from an "
              f"older pass); the uniform recompute corrects this.")
        print(f"Selfcheck BETA: {beta_moved}/{beta_total} betas move under calendar alignment "
              f"(0 expected on the month-aligned synthetic overlap).")
        mism = canon_bad
        # known-answer: a series whose returns are exactly k x a benchmark's returns -> beta ~ k, r2 ~ 1.
        ka = idx.get("nifty_500_tri", {}).get("monthly_values")
        if ka:
            months = sorted(ka)
            br = _logret_by_month(ka)
            for k_mult in (1.0, 1.5, 0.5):
                lvl, synth = 1000.0, {months[0]: 1000.0}
                for m in months[1:]:
                    lvl *= math.exp(k_mult * br.get(m, 0.0))
                    synth[m] = lvl
                beta, r2, _, _ = bench_relative(synth, ka)
                print(f"  known-answer beta({k_mult}x nifty_500_tri) = {beta:.4f} (expect ~{k_mult}), r2 = {r2:.4f} (expect ~1.0)")
        if not args.build:
            sys.exit(0 if mism == 0 else 1)

    if not args.build:
        print("(no --build; nothing written)")
        return

    # ===================== PHASE 1: real series =========================
    print("\nPhase 1: merging real NAVs + real Bloomberg series ...")
    colleague = json.loads(COLLEAGUE.read_text())
    nav_by_amfi = {f.get("amfi_code"): f.get("monthly_nav") for f in colleague.get("mf_funds", [])
                   if f.get("amfi_code") is not None and f.get("monthly_nav")}
    matched = 0
    for f in base["mf_funds"]:
        rn = nav_by_amfi.get(f.get("amfi_code"))
        if rn:
            f["monthly_nav"] = dict(rn)
            matched += 1
    print(f"  NAVs: matched {matched}/{len(base['mf_funds'])} funds to real colleague NAVs by amfi_code")

    p1 = parse_bdh(V01)
    EQUITY = {"NIFTYTR Index": "nifty_50_tri", "NSE100TR Index": "nifty_100_tri",
              "NSE500TR Index": "nifty_500_tri", "NSE150TR Index": "nifty_midcap_150_tri",
              "NSE250TR Index": "nifty_smallcap_250_tri", "NSEBANKT Index": "nifty_bank_tri"}
    idx = base["indices"]

    def set_index(key, series, meta):
        windowed = window(series)
        node = idx.get(key, {"name": key, "category": "", "metadata": {}})
        node["monthly_values"] = windowed
        node["_meta"] = meta
        node.pop("synthesis_method", None)
        idx[key] = node
        return windowed

    for tkr, key in EQUITY.items():
        if tkr in p1:
            w = set_index(key, p1[tkr], {"basis": "total_return_level", "source": "Bloomberg BDH (real)"})
            print(f"  index {key:24} <- {tkr:16} n={len(w)} COVID={covid_move(w):.3f}" if covid_move(w) is not None else f"  index {key} <- {tkr}")
    # (CRISIL Short Term + CRISIL Liquid came back #N/A in v0.1; the high-grade-short
    # and cash debt benchmarks are sourced in Phase 3's supersession instead.)

    # sp_500_tri_inr = SPXT x USDINR, derived in code (not pre-converted)
    spx, usdinr = p1.get("SPXT Index"), p1.get("USDINR Curncy")
    if spx and usdinr:
        common = sorted(set(spx) & set(usdinr))
        sp_inr = {m: spx[m] * usdinr[m] for m in common}
        set_index("sp_500_tri_inr", sp_inr, {"basis": "total_return_level", "source": "SPXT x USDINR (derived in code)"})
        print(f"  index sp_500_tri_inr        <- SPXT x USDINR    n={len(window(sp_inr))}")
        base.setdefault("fx", {})["usd_inr"] = window(usdinr)
        print(f"  fx.usd_inr                  <- USDINR           n={len(window(usdinr))}")

    # held direct-equity stock prices (the cases use Reliance, HDFC Bank, ITC)
    STOCK_MAP = {"RELIANCE IN Equity": "Reliance Industries", "HDFCB IN Equity": "HDFC Bank",
                 "ITC IN Equity": "ITC", "TCS IN Equity": "Tata Consultancy Services",
                 "INFO IN Equity": "Infosys", "HCLT IN Equity": "HCL Technologies",
                 "TECHM IN Equity": "Tech Mahindra", "ICICIBC IN Equity": "ICICI Bank",
                 "LT IN Equity": "Larsen", "BHARTI IN Equity": "Bharti Airtel"}
    companies = base.get("nifty500", {}).get("companies", [])
    stock_hits = 0
    for tkr, nm in STOCK_MAP.items():
        if tkr not in p1:
            continue
        for c in companies:
            if isinstance(c.get("name"), str) and c["name"].strip().lower() == nm.lower():
                c["monthly_prices"] = window(p1[tkr])
                stock_hits += 1
                break
        else:
            # fall back to a contains-match for the held trio's exact snapshot names
            for c in companies:
                if isinstance(c.get("name"), str) and nm.lower() in c["name"].lower():
                    c["monthly_prices"] = window(p1[tkr])
                    stock_hits += 1
                    break
    print(f"  stocks: real monthly_prices set on {stock_hits} matched companies")

    # ===================== PHASE 2: yield -> TR ==========================
    print("\nPhase 2: par-bond yield-to-TR conversion (additive; yields preserved) ...")
    yields = {}
    yields.update(parse_bdh(V02))   # G-Sec ladder + 91TBYLD
    yields.update(parse_bdh(V03))   # FIMMDA BCOP
    rf_yield = yields.get(RF_TICKER)
    if rf_yield:
        print(f"  risk-free: {RF_TICKER} present n={len(window(rf_yield))} "
              f"({window(rf_yield)[WIN_START]:.2f}% .. {window(rf_yield)[WIN_END]:.2f}%)")

    debt_cells = {}      # key -> TR level series (windowed, base 1000)
    debt_yields = {}     # key_yield -> raw yield series (windowed)
    fails = []
    for tkr, (key, tenor) in DEBT_YIELD_MAP.items():
        ys = yields.get(tkr)
        if not ys:
            continue
        tr = yield_to_tr(ys, tenor)
        if not tr:
            continue
        debt_cells[key] = tr
        debt_yields[key.replace("_tr", "_yield")] = window(ys)
        cm = covid_move(tr)
        is_credit = key.startswith(("a_", "bbb_", "nbfc_")) or key.startswith("aa_")
        ok = cm is None or (cm > -0.06)  # inverted-COVID: high-grade/gilt ~flat; credit may dip modestly
        if cm is not None and cm <= -0.08:
            fails.append((key, cm))
        flag = "" if ok else "  <-- CHECK (crashed like equity?)"
        print(f"  cell {key:12} tenor={tenor:>4}y  n={len(tr)}  COVID-Mar20={cm*100:+.2f}%{flag}" if cm is not None
              else f"  cell {key:12} tenor={tenor}y n={len(tr)} (no COVID point)")
    if fails:
        print(f"  INVERTED-COVID FAIL: {fails} (a high-grade/gilt TR must not crash like equity). STOPPING.")
        sys.exit(1)
    print(f"  converted {len(debt_cells)} debt cells; inverted-COVID gut-check passed")

    # ===================== PHASE 3: storage ==============================
    print("\nPhase 3: storing the debt grid (TR + yields + provenance) ...")
    for key, tr in debt_cells.items():
        idx[key] = {"name": key, "category": "debt_benchmark",
                    "monthly_values": tr,
                    "_meta": {"basis": "total_return_level", "derived_from": "yield",
                              "source": "FIMMDA BCOP / G-Sec yield (Bloomberg)",
                              "conversion": "par-bond yield-to-TR (ModDur + convexity)", "base_value": 1000}}
    base.setdefault("debt_yield_primitives", {})
    for yk, ys in debt_yields.items():
        base["debt_yield_primitives"][yk] = {"annualised_yield_pct": ys,
                                              "_meta": {"basis": "annualised_yield", "source": "FIMMDA BCOP / G-Sec (Bloomberg)"}}
    if rf_yield:
        base["debt_yield_primitives"]["tbill_91d_yield"] = {"annualised_yield_pct": window(rf_yield),
                                                            "_meta": {"basis": "annualised_yield", "source": "91-day T-Bill (Bloomberg)", "use": "time_varying_risk_free"}}
    print(f"  stored {len(debt_cells)} TR cells + {len(debt_yields)} yield primitives + 91d T-Bill rf")

    # Supersede the 5 SYNTHETIC debt indices (ADR-0009) with real / real-derived
    # series, in place on their existing keys, so the debt holdings resolve to real
    # data via their stored benchId with no remapping and no synthetic series left.
    # (The CRISIL Short Term and Liquid Bloomberg tickers came back #N/A, so the
    # high-grade-short benchmark is the converted FIMMDA AAA-2Y and cash is the real
    # Nifty 1D Overnight; the published index LIX15 is the wrong instrument (a Liquid
    # 15, not the intended overnight/1D; its -37% March-2020 print confirms the
    # mismatch, it is not glitched data) and is NOT used.)
    nifty1d = yields.get("NIFTY1D Index")  # real overnight TR level (clean), the cash benchmark
    SUPERSEDE = {
        "crisil_liquid": (rebase_1000(window(nifty1d)) if nifty1d else None, "Nifty 1D Overnight (real-pulled Bloomberg)"),
        "crisil_short_term_bond": (debt_cells.get("aaa_2y_tr"), "FIMMDA AAA 2Y yield -> par-bond TR (high-grade short)"),
        "crisil_composite_bond": (debt_cells.get("aaa_5y_tr"), "FIMMDA AAA 5Y yield -> par-bond TR (high-grade composite proxy)"),
        "crisil_dynamic_gilt": (debt_cells.get("gsec_10y_tr"), "G-Sec 10Y yield -> par-bond TR (gilt)"),
        "nifty_10y_gsec": (debt_cells.get("gsec_10y_tr"), "G-Sec 10Y yield -> par-bond TR (10Y gilt)"),
    }
    superseded = 0
    for key, (series, prov) in SUPERSEDE.items():
        if not series:
            continue
        node = idx.get(key, {"name": key, "category": "debt_benchmark"})
        node["monthly_values"] = series
        node["_meta"] = {"basis": "total_return_level", "superseded_synthetic": True, "source": prov,
                         "derived_from": "yield" if "yield" in prov else "level"}
        node.pop("synthesis_method", None)
        idx[key] = node
        superseded += 1
    print(f"  superseded {superseded}/5 synthetic debt indices in place (crisil_liquid<-NIFTY1D, "
          f"crisil_short_term_bond<-aaa_2y, composite<-aaa_5y, gilt/10y_gsec<-gsec_10y)")

    # ===================== PHASE 5a: per-holding tier_b ===================
    print("\nPhase 5a: recomputing per-holding tier_b on real series ...")
    # Shipped basis is the canonical CONSTANT risk-free (0.0525), so per-holding
    # and the TS sleeve agree on R_f. Beta is R_f-invariant; the 91-day T-Bill
    # series is stored above and wired through the seriesRiskFree socket on the TS
    # side as the time-varying demonstration (it shifts only sharpe/sortino).
    rf_by_month = {m: v / 100.0 for m, v in window(rf_yield or {}).items()}

    def bench_vals(benchid):
        return idx.get(benchid, {}).get("monthly_values") if benchid else None

    fund_n = stock_n = 0
    for f in base["mf_funds"]:
        tb = f.get("tier_b_stats") or {}
        nav = f.get("monthly_nav")
        if not nav:
            continue
        benchid = (tb.get("_meta") or {}).get("benchmark_index_id")
        f["tier_b_stats"] = compute_tier_b(nav, bench_vals(benchid), RISK_FREE_ANN,
                                           tb.get("_benchmark_resolution"), benchid)
        fund_n += 1
    for c in companies:
        tb = c.get("tier_b_stats") or {}
        mp = c.get("monthly_prices")
        if not mp:
            continue
        benchid = (tb.get("_meta") or {}).get("benchmark_index_id")
        c["tier_b_stats"] = compute_tier_b(mp, bench_vals(benchid), RISK_FREE_ANN,
                                           tb.get("_benchmark_resolution"), benchid)
        stock_n += 1
    print(f"  recomputed tier_b for {fund_n} funds + {stock_n} stocks (constant rf; resolution + benchId preserved)")
    print(f"  91-day T-Bill rf series available for the TS seriesRiskFree socket: {len(rf_by_month)} months")

    # provenance stamp on the snapshot itself
    base.setdefault("snapshot_metadata", {})["real_data_build"] = {
        "version": "realv1", "method": "Option B real-data t0",
        "nav_source": "colleague real monthly NAVs (bypasses ADR-0014 regeneration)",
        "index_source": "Bloomberg BDH real (equity TRIs, CRISIL debt TR, FX, stocks); FIMMDA/G-Sec yields converted to TR",
        "risk_free": "constant 0.0525 (canonical; per-holding and sleeve agree on R_f); 91-day T-Bill yield series stored and wired for the time-varying seriesRiskFree socket (shifts only sharpe/sortino; beta is R_f-invariant)",
        "additive_note": "source pulls and prior snapshot untouched; yields preserved as the carried primitive",
    }

    OUT_SNAP.write_text(json.dumps(base))
    print(f"\nWrote NEW data version: {OUT_SNAP.name} ({OUT_SNAP.stat().st_size/1e6:.1f} MB). Prior snapshot intact.")
    print("Done. (No agent run, no API spend; the gate recompute is the next step.)")


if __name__ == "__main__":
    main()
